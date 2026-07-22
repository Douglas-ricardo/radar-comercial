# app/workers/notification_tasks.py
import io
import logging
from datetime import datetime

from app.core.celery_app import celery_app
from app.core.clock import utcnow
from app.infrastructure.database import SessionLocal
from app.domain.models import Company, ComputedInsights, NotificationPreference, ScheduledReport, User
from app.services.notification_service import NotificationService
from app.services.live_recency import refresh_days_inactive, company_dataset_max

logger = logging.getLogger(__name__)


@celery_app.task(name="send_daily_notifications", bind=True, max_retries=2)
def send_daily_notifications(self):
    """
    Daily digest dispatched by Celery Beat at 08:00 BRT (11:00 UTC).
    Sends email and/or WhatsApp to users with notifications enabled.
    """
    logger.info("notifications.task.start")
    db = SessionLocal()
    sent = 0
    errors = 0

    try:
        prefs = (
            db.query(NotificationPreference)
            .filter(NotificationPreference.enabled == True)
            .all()
        )

        for pref in prefs:
            try:
                user = db.query(User).filter(
                    User.id == pref.user_id,
                    User.status == "active",
                ).first()

                if not user or user.role not in ("admin", "analyst"):
                    continue

                company = db.query(Company).filter(Company.id == pref.company_id).first()
                if not company:
                    continue

                insights = db.query(ComputedInsights).filter_by(
                    company_id=pref.company_id, date_range="1m"
                ).first()

                if not insights or not insights.opportunities:
                    continue

                # Recência viva (gated por frescor) antes de montar o digest — o
                # email/WhatsApp ao vendedor mostra "dias sem comprar" atualizado.
                dataset_max = company_dataset_max(db, pref.company_id)
                opportunities = [
                    opp for opp in refresh_days_inactive(insights.opportunities, dataset_max)
                    if opp.get("expectedValue", 0) >= pref.min_opportunity_value
                ]

                if not opportunities:
                    continue

                if pref.email_enabled:
                    subject = (
                        f"[Radar Comercial] {len(opportunities)} oportunidades de recuperação"
                        f" — {company.name}"
                    )
                    html = NotificationService.format_opportunity_email(
                        user.name, opportunities, company.name, company.currency
                    )
                    if NotificationService.send_email(user.email, subject, html):
                        sent += 1

                if pref.whatsapp_enabled and pref.whatsapp_phone:
                    msg = NotificationService.format_opportunity_whatsapp(
                        user.name, opportunities, company.currency
                    )
                    NotificationService.send_whatsapp(pref.whatsapp_phone, msg)

            except Exception as exc:
                errors += 1
                logger.error(
                    "notifications.task.user_error",
                    extra={"user_id": pref.user_id, "error": str(exc)},
                    exc_info=True,
                )

        logger.info(
            "notifications.task.complete",
            extra={"sent": sent, "errors": errors, "total_prefs": len(prefs)},
        )
        return {"sent": sent, "errors": errors}

    except Exception as exc:
        logger.error("notifications.task.fatal", extra={"error": str(exc)}, exc_info=True)
        raise self.retry(exc=exc, countdown=300)
    finally:
        db.close()


@celery_app.task(name="send_scheduled_reports", bind=True, max_retries=2)
def send_scheduled_reports(self):
    """
    Celery Beat diário (07:00 UTC = 04:00 BRT). Verifica quais relatórios agendados
    vencem hoje e envia o Excel por email via Resend.
    """
    logger.info("reports.scheduled.task.start")
    db = SessionLocal()
    sent = 0
    errors = 0

    try:
        now_utc = utcnow()
        today_weekday = now_utc.weekday()  # 0=Seg … 6=Dom

        schedules = db.query(ScheduledReport).filter_by(enabled=True).all()

        for sched in schedules:
            try:
                # Verifica se é dia de enviar
                if sched.frequency == "weekly":
                    if sched.day_of_week is None or sched.day_of_week != today_weekday:
                        continue
                elif sched.frequency == "monthly":
                    if now_utc.day != 1:  # sempre dia 1 do mês
                        continue

                # Evita re-envio no mesmo dia
                if sched.last_sent_at and sched.last_sent_at.date() == now_utc.date():
                    continue

                company = db.query(Company).filter_by(id=sched.company_id).first()
                if not company:
                    continue

                insights = db.query(ComputedInsights).filter_by(
                    company_id=sched.company_id, date_range=sched.date_range
                ).first()
                if not insights:
                    continue

                xlsx_bytes = _build_excel(insights, sched.date_range)
                subject = f"[Radar Comercial] Relatório {sched.date_range} — {company.name}"
                html = (
                    f"<p>Olá! Segue em anexo o relatório de oportunidades comerciais "
                    f"({sched.date_range}) de <strong>{company.name}</strong>.</p>"
                    f"<p>— Radar Comercial</p>"
                )

                for recipient in (sched.recipients or []):
                    if NotificationService.send_email_with_attachment(
                        to=recipient,
                        subject=subject,
                        html=html,
                        attachment_bytes=xlsx_bytes,
                        attachment_name=f"radar-{sched.date_range}.xlsx",
                        attachment_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    ):
                        sent += 1

                sched.last_sent_at = now_utc
                db.commit()

            except Exception as exc:
                errors += 1
                logger.error(
                    "reports.scheduled.send_error",
                    extra={"schedule_id": sched.id, "error": str(exc)},
                    exc_info=True,
                )

        logger.info("reports.scheduled.task.complete", extra={"sent": sent, "errors": errors})
        return {"sent": sent, "errors": errors}

    except Exception as exc:
        logger.error("reports.scheduled.task.fatal", extra={"error": str(exc)}, exc_info=True)
        raise self.retry(exc=exc, countdown=300)
    finally:
        db.close()


def _build_excel(insights: ComputedInsights, date_range: str) -> bytes:
    """Constrói o Excel (igual ao endpoint /excel) e retorna bytes para o email."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(color="FFFFFF", bold=True)

    def _hdr(ws, cols):
        for i, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=i, value=col)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")

    def _fit(ws):
        for col_cells in ws.columns:
            mx = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(mx + 4, 40)

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Resumo"
    _hdr(ws_sum, ["Indicador", "Valor"])
    summary = insights.summary or {}
    kpis = [
        ("Receita total (R$)", summary.get("totalRevenue", 0)),
        ("Receita perdida (R$)", summary.get("lostRevenue", 0)),
        ("Taxa de perda (%)", summary.get("lostRate", 0)),
        ("Crescimento (%)", summary.get("revenueGrowth", 0)),
        ("Clientes únicos", summary.get("uniqueCustomers", 0)),
        ("Oportunidades", len(insights.opportunities or [])),
        ("Período", date_range),
    ]
    for r, (label, val) in enumerate(kpis, 2):
        ws_sum.cell(row=r, column=1, value=label)
        ws_sum.cell(row=r, column=2, value=val)
    _fit(ws_sum)

    ws_opp = wb.create_sheet("Oportunidades")
    _hdr(ws_opp, ["Cliente", "Produto", "Valor Esperado (R$)", "Dias Inativo", "Confiança"])
    for r, opp in enumerate(insights.opportunities or [], 2):
        ws_opp.cell(row=r, column=1, value=opp.get("customer", ""))
        ws_opp.cell(row=r, column=2, value=opp.get("product"))
        ws_opp.cell(row=r, column=3, value=opp.get("expectedValue"))
        ws_opp.cell(row=r, column=4, value=opp.get("daysInactive"))
        ws_opp.cell(row=r, column=5, value=opp.get("confidence"))
    _fit(ws_opp)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

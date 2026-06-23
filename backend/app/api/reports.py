# app/api/reports.py
"""Exportação de dados em Excel (.xlsx) — usa openpyxl (já em requirements.txt)."""
from __future__ import annotations

import io
import logging

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.core.auth import get_current_user_and_company
from app.domain.models import ComputedInsights, CustomerProfile
from app.infrastructure.database import get_db_session

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/reports", tags=["Reports"])

_VALID_RANGES = {"1m", "3m", "6m", "12m"}
_HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _header_row(ws, cols: list[str]) -> None:
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 20


def _autofit(ws) -> None:
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 40)


@router.get("/{company_id}/excel")
def export_excel(
    company_id: str,
    date_range: str = Query("6m"),
    branch: str | None = Query(default=None),
    salesperson: str | None = Query(default=None),
    token_data=Depends(get_current_user_and_company),
    db: Session = Depends(get_db_session),
):
    """Exporta oportunidades em .xlsx com abas Resumo e Oportunidades."""
    if token_data.company_id != company_id:
        raise HTTPException(status_code=403, detail="Acesso negado.")
    if date_range not in _VALID_RANGES:
        raise HTTPException(status_code=400, detail=f"Período inválido: {date_range}.")

    # Scope automático (mesmo mecanismo da carteira)
    effective_branch = branch
    if token_data.scope and token_data.role != "admin":
        parts = token_data.scope.split(":", 1)
        if parts[0] == "branch" and len(parts) == 2:
            effective_branch = parts[1]

    row = db.query(ComputedInsights).filter_by(company_id=company_id, date_range=date_range).first()
    if not row:
        raise HTTPException(status_code=404, detail="Nenhum dado para o período. Faça upload primeiro.")

    summary = row.summary or {}
    raw_opps = row.opportunities or []

    # Batch-join com CustomerProfile para branch/salesperson/document_id
    hashes = [opp.get("customerHash", "") for opp in raw_opps if opp.get("customerHash")]
    profiles_q = db.query(
        CustomerProfile.customer_hash,
        CustomerProfile.branch,
        CustomerProfile.salesperson,
        CustomerProfile.document_id,
    ).filter(
        CustomerProfile.company_id == company_id,
        CustomerProfile.customer_hash.in_(hashes),
    )
    if effective_branch:
        profiles_q = profiles_q.filter(CustomerProfile.branch == effective_branch)
    if salesperson:
        profiles_q = profiles_q.filter(CustomerProfile.salesperson == salesperson)
    profile_map = {r[0]: {"branch": r[1], "salesperson": r[2], "document_id": r[3]} for r in profiles_q.all()}

    # Filtrar oportunidades pelo scope/filters
    scope_hashes = set(profile_map.keys()) if (effective_branch or salesperson) else None
    opps = [
        opp for opp in raw_opps
        if scope_hashes is None or opp.get("customerHash", "") in scope_hashes
    ]

    wb = Workbook()

    # ── Aba 1: Resumo ──────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Resumo"
    _header_row(ws_sum, ["Indicador", "Valor"])
    kpis = [
        ("Receita total (R$)", summary.get("totalRevenue", 0)),
        ("Receita perdida (R$)", summary.get("lostRevenue", 0)),
        ("Taxa de perda (%)", summary.get("lostRate", 0)),
        ("Crescimento de receita (%)", summary.get("revenueGrowth", 0)),
        ("Clientes únicos", summary.get("uniqueCustomers", 0)),
        ("Produtos únicos", summary.get("uniqueProducts", 0)),
        ("Oportunidades exportadas", len(opps)),
        ("Período", date_range),
    ]
    for r, (label, val) in enumerate(kpis, 2):
        ws_sum.cell(row=r, column=1, value=label)
        ws_sum.cell(row=r, column=2, value=val)
    _autofit(ws_sum)

    # ── Aba 2: Oportunidades ───────────────────────────────────────────────────
    ws_opp = wb.create_sheet("Oportunidades")
    opp_cols = [
        "Cliente", "CNPJ/CPF", "Filial", "Vendedor",
        "Produto", "Valor Esperado (R$)", "Dias Inativo",
        "Frequência", "Confiança", "Segmento",
    ]
    _header_row(ws_opp, opp_cols)
    for r, opp in enumerate(opps, 2):
        ch = opp.get("customerHash", "")
        prof = profile_map.get(ch, {})
        ws_opp.cell(row=r, column=1, value=opp.get("customer", opp.get("customerName", "")))
        ws_opp.cell(row=r, column=2, value=prof.get("document_id"))
        ws_opp.cell(row=r, column=3, value=prof.get("branch"))
        ws_opp.cell(row=r, column=4, value=prof.get("salesperson"))
        ws_opp.cell(row=r, column=5, value=opp.get("product"))
        ws_opp.cell(row=r, column=6, value=opp.get("expectedValue"))
        ws_opp.cell(row=r, column=7, value=opp.get("daysInactive"))
        ws_opp.cell(row=r, column=8, value=opp.get("frequency"))
        ws_opp.cell(row=r, column=9, value=opp.get("confidence"))
        ws_opp.cell(row=r, column=10, value=opp.get("type"))
    _autofit(ws_opp)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    logger.info("reports.excel.exported", extra={
        "company_id": company_id, "date_range": date_range, "rows": len(opps),
    })

    return StreamingResponse(
        content=buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="radar-{date_range}.xlsx"'},
    )
